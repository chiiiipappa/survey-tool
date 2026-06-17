"""
STEP3 ファン度分析結果を openpyxl でExcelワークブックに変換する。

main.py 等の集計エンジンを再計算せず、フロントエンドが保持している
FanAnalysisResponse（直前の分析結果そのもの）をそのままシート化する。
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import FanAnalysisResponse

_HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_LABEL_FONT   = Font(bold=True)
_SECTION_FONT = Font(bold=True, size=11)


def _header(ws, row: int, col: int, value, width: float | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _bold(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _LABEL_FONT
    return cell


def _section(ws, row: int, value):
    cell = ws.cell(row=row, column=1, value=value)
    cell.font = _SECTION_FONT
    return cell


# ---------------------------------------------------------------------------
# シート1: ファン度判定結果（回答者単位）
# ---------------------------------------------------------------------------

def _write_results_sheet(wb: Workbook, body: FanAnalysisResponse):
    ws = wb.create_sheet(title="判定結果")
    headers = [
        "response_id", body.row_question_text or "縦軸回答", body.col_question_text or "横軸回答",
        "fan_degree_label", "判定status", "is_core_fan", "is_fan_or_above",
        "is_light_fan_or_above", "is_fan_degree_valid",
    ]
    widths = [12, 36, 36, 16, 12, 12, 14, 18, 16]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        _header(ws, 1, ci, h, width=w)

    for ri, row in enumerate(body.respondent_rows, start=2):
        ws.cell(row=ri, column=1, value=row.response_id)
        ws.cell(row=ri, column=2, value=row.row_answer)
        ws.cell(row=ri, column=3, value=row.col_answer)
        ws.cell(row=ri, column=4, value=row.fan_degree_label)
        ws.cell(row=ri, column=5, value=row.status)
        ws.cell(row=ri, column=6, value=row.is_core_fan)
        ws.cell(row=ri, column=7, value=row.is_fan_or_above)
        ws.cell(row=ri, column=8, value=row.is_light_fan_or_above)
        ws.cell(row=ri, column=9, value=row.is_fan_degree_valid)


# ---------------------------------------------------------------------------
# シート2: ファン度マトリクス
# ---------------------------------------------------------------------------

def _write_matrix_sheet(wb: Workbook, body: FanAnalysisResponse):
    ws = wb.create_sheet(title="マトリクス")
    headers = [body.row_question_text or "行選択肢", body.col_question_text or "列選択肢", "fan_degree_label"]
    for ci, h in enumerate(headers, start=1):
        _header(ws, 1, ci, h, width=34)

    ri = 2
    for cell in body.matrix:
        if not cell.label:
            continue
        ws.cell(row=ri, column=1, value=cell.row_value)
        ws.cell(row=ri, column=2, value=cell.col_value)
        ws.cell(row=ri, column=3, value=cell.label)
        ri += 1


# ---------------------------------------------------------------------------
# シート3: ファン度集計表
# ---------------------------------------------------------------------------

def _write_summary_sheet(wb: Workbook, body: FanAnalysisResponse):
    ws = wb.create_sheet(title="集計表")
    s = body.summary

    _header(ws, 1, 1, "ファン度", width=16)
    _header(ws, 1, 2, "N", width=10)
    _header(ws, 1, 3, "%", width=10)
    _header(ws, 1, 4, "累積%", width=10)

    ri = 2
    for c in s.counts:
        ws.cell(row=ri, column=1, value=c.label)
        ws.cell(row=ri, column=2, value=c.n)
        ws.cell(row=ri, column=3, value=round(c.pct, 1)).number_format = "0.0"
        ws.cell(row=ri, column=4, value=round(c.cum_pct, 1)).number_format = "0.0"
        ri += 1

    ri += 1
    _section(ws, ri, "【サマリー】")
    ri += 1
    metrics = [
        ("分母（人数）", s.denominator_n),
        ("分母の種類", s.denominator_mode),
        ("コアファン率(%)", round(s.core_fan_rate, 1)),
        ("ファン以上率(%)", round(s.fan_or_above_rate, 1)),
        ("ライトファン以上率(%)", round(s.light_fan_or_above_rate, 1)),
        ("判定不能数", s.undetermined_n),
        ("除外数", s.excluded_n),
    ]
    for label, value in metrics:
        _bold(ws, ri, 1, label)
        ws.cell(row=ri, column=2, value=value)
        ri += 1


# ---------------------------------------------------------------------------
# ワークブック組み立て
# ---------------------------------------------------------------------------

def build_fan_excel_workbook(body: FanAnalysisResponse) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    _write_results_sheet(wb, body)
    _write_matrix_sheet(wb, body)
    _write_summary_sheet(wb, body)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
