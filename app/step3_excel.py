"""
STEP3 クロス集計結果を openpyxl でExcelワークブックに変換する。
各設問を1シートに出力し、%表・N表・編集可能なネイティブグラフを含む。
"""
import io
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import ExportQuestion, Step3ExportRequest

# ---------------------------------------------------------------------------
# チャートタイプ表示名
# ---------------------------------------------------------------------------
_CHART_TYPE_LABELS = {
    "bar":        "棒グラフ",
    "grouped":    "grouped棒",
    "stacked100": "100%積み上げ",
    "pie":        "円グラフ",
    "avg_bar":    "平均棒",
    "table_only": "表のみ",
}
_ORIENT_LABELS = {"v": "（縦）", "h": "（横）"}


def _chart_type_label(q: ExportQuestion) -> str:
    base = _CHART_TYPE_LABELS.get(q.chart_type, q.chart_type)
    if q.chart_type in ("bar", "grouped", "stacked100"):
        base += _ORIENT_LABELS.get(q.orientation, "")
    return base


# ---------------------------------------------------------------------------
# スタイルヘルパー
# ---------------------------------------------------------------------------
_HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_LABEL_FONT   = Font(bold=True)
_SECTION_FONT = Font(bold=True, size=11)


def _header(ws, row: int, col: int, value, width: float = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _bold(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _LABEL_FONT
    return cell


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# シート書き込み
# ---------------------------------------------------------------------------

def _write_question_sheet(
    wb: Workbook,
    q: ExportQuestion,
    axis_question_text: str,
    axis_categories: List[str],
    axis_totals: List[int],
):
    title = q.graph_title or q.question_text
    sheet_name = q.question_code[:31]  # Excelシート名は31文字制限
    ws = wb.create_sheet(title=sheet_name)

    n_cats = len(axis_categories)

    # ────────────────────────────────────────────
    # メタ情報（A1〜A4）
    # ────────────────────────────────────────────
    meta = [
        ("設問コード",  q.question_code),
        ("質問文",      title),
        ("集計軸",      axis_question_text),
        ("グラフ種類",  _chart_type_label(q)),
    ]
    for i, (label, value) in enumerate(meta):
        row = i + 1
        _bold(ws, row, 1, label)
        ws.cell(row=row, column=2, value=value)
    _set_col_width(ws, 1, 14)
    _set_col_width(ws, 2, 40)

    # ────────────────────────────────────────────
    # %表（行6〜）
    # ────────────────────────────────────────────
    PCT_SECTION_ROW = 6
    PCT_HEADER_ROW  = 7
    pct_data_start  = 8
    pct_data_end    = 7 + len(q.rows)

    ws.cell(row=PCT_SECTION_ROW, column=1, value="【%表】").font = _SECTION_FONT

    # ヘッダー
    _header(ws, PCT_HEADER_ROW, 1, "選択肢", width=30)
    for ci, (cat, tot) in enumerate(zip(axis_categories, axis_totals)):
        _header(ws, PCT_HEADER_ROW, ci + 2, f"{cat}\n(n={tot})", width=14)

    # データ行
    for ri, row_data in enumerate(q.rows):
        r = pct_data_start + ri
        ws.cell(row=r, column=1, value=row_data.label)
        for ci, pct in enumerate(row_data.percents):
            cell = ws.cell(row=r, column=ci + 2, value=round(pct, 1))
            cell.number_format = "0.0"

    # ────────────────────────────────────────────
    # N表（pct_data_end + 2〜）
    # ────────────────────────────────────────────
    N_SECTION_ROW = pct_data_end + 2
    N_HEADER_ROW  = pct_data_end + 3
    n_data_start  = pct_data_end + 4

    ws.cell(row=N_SECTION_ROW, column=1, value="【N表】").font = _SECTION_FONT

    _header(ws, N_HEADER_ROW, 1, "選択肢")
    for ci, (cat, tot) in enumerate(zip(axis_categories, axis_totals)):
        _header(ws, N_HEADER_ROW, ci + 2, f"{cat}\n(n={tot})")

    for ri, row_data in enumerate(q.rows):
        r = n_data_start + ri
        ws.cell(row=r, column=1, value=row_data.label)
        for ci, cnt in enumerate(row_data.counts):
            ws.cell(row=r, column=ci + 2, value=cnt)

    # ────────────────────────────────────────────
    # グラフ（%表データを参照）
    # ────────────────────────────────────────────
    if q.chart_type == "table_only" or not q.rows:
        return

    # グラフのアンカー: N表の下か右側（列数が少ない場合は右に配置）
    chart_anchor_col = n_cats + 3
    chart_anchor = f"{get_column_letter(chart_anchor_col)}1"

    if q.chart_type == "pie":
        charts = _build_pie_charts(ws, q, PCT_HEADER_ROW, pct_data_end, n_cats, axis_categories)
        for ci, chart in enumerate(charts):
            offset_col = chart_anchor_col + ci * 8
            ws.add_chart(chart, f"{get_column_letter(offset_col)}1")
    else:
        chart = _build_bar_chart(ws, q, PCT_HEADER_ROW, pct_data_end, n_cats, title)
        ws.add_chart(chart, chart_anchor)


# ---------------------------------------------------------------------------
# BarChart（bar / grouped / stacked100 / avg_bar）
# ---------------------------------------------------------------------------

def _build_bar_chart(
    ws,
    q: ExportQuestion,
    header_row: int,
    data_end_row: int,
    n_cats: int,
    title: str,
) -> BarChart:
    chart = BarChart()
    chart.title = title

    if q.chart_type == "stacked100":
        chart.grouping = "percentStacked"
        chart.overlap  = 100
    else:
        chart.grouping = "clustered"

    chart.type = "col" if q.orientation == "v" else "bar"

    # データ参照（%表のヘッダー行含む）
    data = Reference(
        ws,
        min_col=2, max_col=1 + n_cats,
        min_row=header_row, max_row=data_end_row,
    )
    cats = Reference(
        ws,
        min_col=1,
        min_row=header_row + 1, max_row=data_end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # 系列色適用（resolved_colors = 軸カテゴリー分の色）
    for i, series in enumerate(chart.series):
        if i < len(q.resolved_colors):
            series.graphicalProperties.solidFill = q.resolved_colors[i].lstrip("#")

    chart.width  = 18
    chart.height = 12
    return chart


# ---------------------------------------------------------------------------
# PieChart（軸カテゴリーごとに1グラフ）
# ---------------------------------------------------------------------------

def _build_pie_charts(
    ws,
    q: ExportQuestion,
    header_row: int,
    data_end_row: int,
    n_cats: int,
    axis_categories: List[str],
) -> List[PieChart]:
    charts = []
    n_rows = data_end_row - header_row  # 選択肢数

    for ci in range(n_cats):
        chart = PieChart()
        chart.title = axis_categories[ci] if ci < len(axis_categories) else f"系列{ci+1}"

        # 1列分のデータ（ヘッダー含む）
        data = Reference(
            ws,
            min_col=ci + 2, max_col=ci + 2,
            min_row=header_row, max_row=data_end_row,
        )
        cats = Reference(
            ws,
            min_col=1,
            min_row=header_row + 1, max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # DataPoint で各選択肢に色を付ける
        for ri in range(n_rows):
            if ri < len(q.resolved_colors):
                dp = DataPoint(idx=ri)
                dp.graphicalProperties.solidFill = q.resolved_colors[ri].lstrip("#")
                chart.series[0].dPt.append(dp)

        chart.width  = 10
        chart.height = 10
        charts.append(chart)

    return charts


# ---------------------------------------------------------------------------
# ワークブック組み立て
# ---------------------------------------------------------------------------

def build_excel_workbook(body: Step3ExportRequest) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    for q in body.questions:
        _write_question_sheet(
            wb,
            q,
            body.axis_question_text,
            body.axis_categories,
            body.axis_totals,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
